import openpyxl
import os
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font, colors

# fillColour = PatternFill(patternType='solid', fgColor=colors.Color(rgb='00FF0000'))
# currentSheet1.cell(row=5, column=4).fill = redFill
# code to fill cell in red

class ExcelWB:
    def __init__(self) -> None:
        self.WB = openpyxl.load_workbook(os.curdir +"/INCOMING1.xlsx")
        self.template = self.WB['template']
        self.sheet = None
        self.unfilledColour = PatternFill(patternType='solid', fgColor=colors.Color(rgb='00FFC1C1'))
        self.unabletoextractColour = PatternFill(patternType='solid', fgColor=colors.Color(rgb='00ebff6a'))
        
    def getSheet(self,month, year):
        try:
            sheet = self.WB[month+'-'+year]
        except:
            sheet = self.WB.copy_worksheet(self.template)
            sheet.title = month+'-'+year
            print('created new sheet')
        self.sheet = sheet
    
    def write(self,data,row):
        for i,v in enumerate(data):
            if v == "unfilled":
                self.sheet.cell(row,i+1).fill = self.unfilledColour
            elif v == "unable to extract:":
                self.sheet.cell(row,i+1).fill = self.unabletoextractColour
                self.sheet.cell(row,i+1,value=v)
            else:
                self.sheet.cell(row,i+1,value=v)
    
    def save(self,name):
        self.WB.save(name+'.xlsx')

        

