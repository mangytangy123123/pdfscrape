import pdfreader
from openpyxl import load_workbook
from pdfreader import PDFDocument, SimplePDFViewer
import os
from os import listdir
from os.path import isfile, join



IncomingPDF = 'Electronic funds transfer - Incoming (EFTI)'

def handleBeneficiary(doc, ind, sheet):
    doc.navigate(3)
    doc.render()
    x = doc.canvas.strings
    print(x[x.index("1. Full name of entity:")+2],x[x.index("1. Full name of entity:")+4]," in canada")
    sheet.cell(ind+2,5,value=x[x.index("1. Full name of entity:")+4]+" " +x[x.index("1. Full name of entity:")+2])
    sheet.cell(ind+2,6,value=x[x.index("5. Street address:")+1]+ ' ' + x[x.index("5. Street address:")+9])
    second_page = doc.canvas.strings
    sheet.cell(ind+2,7,value=second_page[second_page.index("11. Date of birth:")+1])
    sheet.cell(ind+2,8,value=second_page[second_page.index("10. Telephone number:")+1])
    sheet.cell(ind+2,9,value=second_page[second_page.index("12. Occupation:")+1])
    sheet.cell(ind+2,10,value=second_page[second_page.index("14. Identifier:")+1])
    sheet.cell(ind+2,11,value='BENEFICIARY') 

def checkID(id, second_page,ind, sheet):
    if id == "Other":
        sheet.cell(ind+2,10,value=second_page[second_page.index("14a. If \'Other\', please specify:" )+1])
        sheet.cell(ind+2,11,value=second_page[second_page.index("15. ID number:")+1])
        return False
    else:
        return True


def incoming(x,ind, doc, sheet):
    sheet.cell(ind+2,1,value=x[x.index("External report number:")+1])
    sheet.cell(ind+2,2,value=x[x.index("External report number:")+3])
    sheet.cell(ind+2,3,value=x[x.index("3. Amount of transaction:")+1])
    sheet.cell(ind+2,4,value=x[x.index("3. Amount of transaction:")+3])
    if x[x.index('7. Country:')+1] != 'Canada':

        handleBeneficiary(doc, ind, sheet)
    else:
        print(x[x.index("1. Full name of entity:")+2],x[x.index("1. Full name of entity:")+4]," in canada")
        sheet.cell(ind+2,5,value=x[x.index("1. Full name of entity:")+4]+" " +x[x.index("1. Full name of entity:")+2])
        sheet.cell(ind+2,6,value=x[x.index("5. Street address:")+1]+" "+x[x.index("5. Street address:")+9])
        doc.navigate(2); doc.render()
        second_page = doc.canvas.strings
        sheet.cell(ind+2,7,value=second_page[second_page.index("11. Date of birth:")+1])
        if "10. Telephone number:" in second_page:
            sheet.cell(ind+2,8,value=second_page[second_page.index("10. Telephone number:")+1])
        else:
            sheet.cell(ind+2,8,value=x[x.index("10. Telephone number:")+1])
        sheet.cell(ind+2,9,value=second_page[second_page.index("12. Occupation:")+1])
        if checkID(second_page[second_page.index("14. Identifier:")+1], second_page, ind, sheet):
            sheet.cell(ind+2,10,value=second_page[second_page.index("14. Identifier:")+1])
            sheet.cell(ind+2,11,value=second_page[second_page.index("15. ID number:")+1])

def loop(onlyfiles,i, sheet):
        pdfFileObj = open(onlyfiles, 'rb')
        doc = SimplePDFViewer(pdfFileObj)
        
        doc.navigate(1)
        doc.render()
        text = doc.canvas.strings
        if IncomingPDF in text:
            incoming(text,i, doc, sheet)

        pdfFileObj.close()

def main():
    wb = load_workbook(filename='INCOMING1.xlsx')
    ind = 0
    for year in os.listdir("C:/Users/lawre.DESKTOP-SCR7VIT/Documents/python/pdfs/"):
        for month in os.listdir("C:/Users/lawre.DESKTOP-SCR7VIT/Documents/python/pdfs/"+year):
            try:
                sheet = wb[month+'-'+year]
            except:
                temp = wb.copy_worksheet(wb['template'])
                temp.title = month+'-'+year
            sheet = wb[month+'-'+year]

            sheet = wb[month+'-'+year]
            for sub in os.listdir("C:/Users/lawre.DESKTOP-SCR7VIT/Documents/python/pdfs/" +year+'/'+month):
                for file in os.listdir("C:/Users/lawre.DESKTOP-SCR7VIT/Documents/python/pdfs/"+year+"/"+month+'/'+sub):
                    dir = "C:/Users/lawre.DESKTOP-SCR7VIT/Documents/python/pdfs/"+year+"/"+month+'/'+sub+'/'+file
                    loop(dir, ind,sheet)
                    ind += 1
            ind = 0

    
    #print(everything)
    wb.save('test.xlsx')
main()

